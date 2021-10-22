from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.shortcuts import render, redirect
from django.urls import reverse

from django.contrib.auth.decorators import login_required
# Create your views here.
from .models import Question, Choice
from configurations.models import Configurations, Department, CoursOfDepartment
#
from .forms import QuestionForm
# để phân trang
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
#
#
import openpyxl
#
@login_required()
def index(request):
    group = request.user.groups.values_list('name',flat = True).first() # QuerySet Object
   
    #for g in request.user.groups.all():
    #    l.append(g.name)
    #return HttpResponse(l)
    config = Configurations.objects.filter(id = 1).first()
    list_departments = []
    departments = Department.objects.all()
    for dep in departments:
        cour =  CoursOfDepartment.objects.filter(department_id =dep.id).first()
        if cour != None:
            list_departments.append(dep)

    courseOfDepartments = CoursOfDepartment.objects.all()
    #return HttpResponse(sections)
    return render(request, 'question/index.html',{
                                                'group':group ,
                                                'config':config,
                                                'departments':list_departments,
                                                'courseOfDepartments':courseOfDepartments,
                                             })

@login_required()
def add_question(request, course_id):
    group = request.user.groups.values_list('name',flat = True).first() # QuerySet Object
                                      # QuerySet to `list`
   
    #for g in request.user.groups.all():
    #    l.append(g.name)
    #return HttpResponse(l)
    config = Configurations.objects.filter(id = 1).first()
    departments = Department.objects.all()
    course = CoursOfDepartment.objects.filter( id = course_id).first()
    if request.method == 'POST':
        choice_texts = None
        question = Question()
        if 'course_id' in request.POST :
            #question_type = request.POST['question_type']
            question.coursOfDepartment_id = request.POST['course_id']
        if 'question_type' in request.POST :
            #question_type = request.POST['question_type']
            question.question_type = request.POST['question_type']
        if 'question_name' in request.POST :
            question.question_name = request.POST['question_name']
        if 'chapter' in request.POST :
            question.chapter = request.POST['chapter']
        if 'question_level' in request.POST:
            question.question_level = request.POST['question_level']
        if 'question_name' in request.POST :
            question.question_name = request.POST['question_name']
        if 'choice_text[]' in request.POST :
            choice_texts = request.POST.getlist("choice_text[]")
            
        if 'correct_answer[]' in request.POST :
            answers = request.POST.getlist("correct_answer[]")
        

        correct_answers = ""
        if choice_texts:
            for i, item in enumerate(choice_texts):

                if 'Choice'+ str(i+1) in answers:
                    correct_answers = correct_answers + item
            question.correct_answer =   correct_answers
        question.save()

        if choice_texts:
            for i, item in enumerate(choice_texts):
                choice = Choice()
                choice.question_id = question.id
                choice.choice_name = item
                choice.save()
            
          
        #return HttpResponse(correct_answers )
    return render(request, 'question/add_question.html',{
                                                'group':group ,
                                                'config':config,
                                                'departments':departments,
                                                'course':course,
                                             })
# Danh sách 
Question_type = [
        ('text', 'Điền vào chổ trống'),
        ('radio', 'Chọn đáp án đúng'),
        ('checkbox', 'Chọn nhiều đáp án'),
        ('textarea', 'Điền vào chổ trống (nhiều dòng)'),
]
Question_level = [
        ('cb', 'Cơ bản'),
        ('nc', 'Nâng cao'),
]
#
@login_required()
def list_question(request, course_id):






    group = request.user.groups.values_list('name',flat = True).first() # QuerySet Object
                                      # QuerySet to `list`
   
    #for g in request.user.groups.all():
    #    l.append(g.name)
    #return HttpResponse(l)
    config = Configurations.objects.filter(id = 1).first()
    departments = Department.objects.all()
    course = CoursOfDepartment.objects.filter( id = course_id).first()
    questions = Question.objects.filter( coursOfDepartment_id = course_id)
    
    #phân trang
    page = request.GET.get('page', 1)
    paginator = Paginator(questions, 20)
    try:
        questions_page = paginator.page(page)
    except PageNotAnInteger:
        questions_page = paginator.page(1)
    except EmptyPage:
        questions_page = paginator.page(paginator.num_pages)
    #if request.method == 'POST':
      
    return render(request, 'question/list_question.html',{
                                                'group':group ,
                                                'config':config,
                                                'departments':departments,
                                                'course':course,
                                                'questions':questions_page,
                                                'Question_type': Question_type,
                                                'Question_level':Question_level
                                             })

@login_required()
def add_question1(request, course_id):
    group = request.user.groups.values_list('name',flat = True).first() # QuerySet Object
                                      # QuerySet to `list`
   
    #for g in request.user.groups.all():
    #    l.append(g.name)
    #return HttpResponse(l)
    config = Configurations.objects.filter(id = 1).first()
    departments = Department.objects.all()
    course = CoursOfDepartment.objects.filter( id = course_id).first()
    if request.method == 'POST':
        question = Question()
        if 'course_id' in request.POST :
            #question_type = request.POST['question_type']
            question.coursOfDepartment_id = request.POST['course_id']
        if 'question_type' in request.POST :
            #question_type = request.POST['question_type']
            question.question_type = request.POST['question_type']
        if 'question_name' in request.POST :
            question.question_name = request.POST['question_name']
        if 'chapter' in request.POST :
            question.chapter = request.POST['chapter']
        if 'question_level' in request.POST:
            question.question_level = request.POST['question_level']
        if 'question_name' in request.POST :
            question.question_name = request.POST['question_name']
        if 'choice_text[]' in request.POST :
            choice_texts = request.POST.getlist("choice_text[]")
            
        if 'correct_answer[]' in request.POST :
            answers = request.POST.getlist("correct_answer[]")
        #return HttpResponse(enumerate(choice_texts) )

        correct_answers = ""
        for i, item in enumerate(choice_texts):

            if 'Choice'+ str(i+1) in answers:
                correct_answers = correct_answers + item

        question.correct_answer =   correct_answers
        question.save()
        
        for i, item in enumerate(choice_texts):
            choice = Choice()
            choice.question_id = question.id
            choice.choice_name = item
            choice.save()
            
          
        #return HttpResponse(correct_answers )
    return render(request, 'question/add_question.html',{
                                                'group':group ,
                                                'config':config,
                                                'departments':departments,
                                                'course':course,
                                             })

@login_required()
def edit_question(request, question_id):
    group = request.user.groups.values_list('name',flat = True).first() # QuerySet Object
                                      # QuerySet to `list`
   
    #for g in request.user.groups.all():
    #    l.append(g.name)
    #return HttpResponse(l)
    config = Configurations.objects.filter(id = 1).first()
    departments = Department.objects.all()
    #course = CoursOfDepartment.objects.filter( id = course_id).first()
    question = Question.objects.filter(id = question_id).first()
    choices = Choice.objects.filter( question_id = question_id )
    if request.method == 'POST':
        #question = Question()
        choice_texts = None
        answers = None
        if 'course_id' in request.POST :
            #question_type = request.POST['question_type']
            question.coursOfDepartment_id = request.POST['course_id']
        if 'question_type' in request.POST :
            #question_type = request.POST['question_type']
            question.question_type = request.POST['question_type']
        if 'question_name' in request.POST :
            question.question_name = request.POST['question_name']
        if 'chapter' in request.POST :
            question.chapter = request.POST['chapter']
        if 'question_level' in request.POST:
            question.question_level = request.POST['question_level']
        if 'question_name' in request.POST :
            question.question_name = request.POST['question_name']
        if 'choice_text[]' in request.POST :
            choice_texts = request.POST.getlist("choice_text[]") 
        if 'correct_answer[]' in request.POST :
            answers = request.POST.getlist("correct_answer[]")
        #else:
            #some_operation(x)
        #if choice_texts in locals():
        #return HttpResponse(answers )
        
        correct_answers = ""
        if choice_texts:
            for i, item in enumerate(choice_texts):

                if 'Choice'+ str(i+1) in answers:
                    correct_answers = correct_answers + item

           
             #// luu danh sách đáp án mới
            for i, item in enumerate(choice_texts):
                choice = Choice()
                choice.question_id = question.id
                choice.choice_name = item
                choice.save()

        #question.save()
        #// luu danh sách đáp án củ
        for choice in choices:
            text = 'choice_text'+ str(choice.id)
            if text in request.POST:
                choice.choice_name = request.POST[ text]
                choice.save()
            
            textChoiceOld ='ChoiceOld'+ str(choice.id)
           
            if textChoiceOld in answers:
                correct_answers = correct_answers + request.POST[ text]
         #// luu danh sách đáp án mới
        question.correct_answer =   correct_answers
        question.save()
    
    course = CoursOfDepartment.objects.filter( id = question.coursOfDepartment_id ).first()
    #
    form = QuestionForm()
  
    #return HttpResponse(form )
    return render(request, 'question/edit_question.html',{
                                                'group':group ,
                                                'config':config,
                                                'departments':departments,
                                                'course':course,
                                                'question':question,
                                                'Question_type': Question_type,
                                                'Question_level':Question_level,
                                                'choices':choices,
                                                'form':form
                                             })

@login_required()
def import_question(request, course_id ):
    if "POST" == request.method:
        
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        #print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        i = 0
        for row in worksheet.iter_rows():
            row_data = list()
            if i!=0:
               
                
                question = Question()
                
                    #question_type = request.POST['question_type']
                question.coursOfDepartment_id = course_id
                
                question.question_type = 'radio'
                question.chapter = row[1].value #request.POST['chapter']
                question.question_name = row[2].value
                question.question_level = row[3].value
                question.correct_answer =   row[4].value

                question.save()
                choice_texts = [4,5,6,7]
               
                if choice_texts:
                    for i, item in enumerate(choice_texts):
                       
                        if(row[item].value != None):
                            choice = Choice()
                            choice.question_id = question.id
                            choice.choice_name = row[item].value
                            choice.save()
                        
            i = i +1

    return redirect('question:list_question', course_id=course_id)
    return HttpResponseRedirect(reverse('question:list_question'))
    return HttpResponseRedirect('question/' + str(course_id )+  '/list_question/')

@login_required()
def import_question_old(request, course_id ):
    if "POST" == request.method:
        
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        #print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        i = 1
        question_old = Question()
        correct_answer=""
        for row in worksheet.iter_rows():
            row_data = list()
           
            if i % 5 ==1:
                
                question = Question()
                question.coursOfDepartment_id = course_id               
                question.question_type = 'radio'
                question.chapter = row[0].value #request.POST['chapter']
                
                question.question_name = row[1].value #row[0].value
                question.question_level = 'cb' #request.POST['question_level'] # row[3].value
                correct_answer = row[2].value #row[1].value
                question.save()
                
                question_old = question
                
            else:
                
                #question = Question.objects.filter(id = question_id).first()
                choice = Choice()
                choice.question_id = question_old.id
                
                choice.choice_name = row[1].value #row[0].value
                choice.save()
               
                if  correct_answer== "A" and (i % 5)==2:
                    question_old.correct_answer = row[1].value #row[0].value
                    question_old.save()
                elif  correct_answer== "B"  and (i % 5)==3:
                    question_old.correct_answer = row[1].value #row[0].value
                    question_old.save()
                elif  correct_answer== "C"  and (i % 5)==4:
                    question_old.correct_answer = row[1].value #row[0].value
                    question_old.save()
                elif  correct_answer== "D" and (i % 5)==0:
                    question_old.correct_answer = row[1].value #row[0].value
                    question_old.save()
            i = i +1

    return redirect('question:list_question', course_id=course_id)
